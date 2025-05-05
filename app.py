import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import re
from werkzeug.security import generate_password_hash, check_password_hash

from flask import Flask, render_template, jsonify, send_file, request
from flask_cors import CORS
import os
from openpyxl import Workbook, load_workbook
import pandas as pd
import logging
from datetime import datetime  # Import datetime for timestamps
import tempfile

app = Flask(__name__, template_folder='.')
CORS(app)  # Enable CORS for all routes
app.secret_key = 'your_secret_key'  # Necessary for flashing messages


# Create folder to store uploaded images
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# creating excel if not present 
REGISTRATION_FILE = 'registrations_data.xlsx'
REPORT_FILE = 'report_potholes.xlsx'
CONTACT_FILE = 'contact_us.xlsx'  # New file for contact form submissions
# Path to your Excel file - update this to the actual path
EXCEL_FILE_PATH = 'report_potholes.xlsx'

# Create Excel files with headers if they don't exist
if not os.path.exists(CONTACT_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Timestamp", "Message"])  # Headers for contact form
    workbook.save(CONTACT_FILE)

if not os.path.exists(REGISTRATION_FILE):
    workbook = Workbook()
    sheet = workbook.active
    # Create header row with a Timestamp column
    sheet.append(["Timestamp", "First Name", "Last Name", "Email", "Phone", "Username", "Password"])
    workbook.save(REGISTRATION_FILE)

if not os.path.exists(REPORT_FILE):
    workbook = Workbook()
    sheet = workbook.active
    # Create header row with a Timestamp column and a Status column
    sheet.append(["Timestamp", "Image Path", "Pin Code", "Latitude", "Longitude", "Status"])
    workbook.save(REPORT_FILE)

#returning registration page
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/register', methods=['POST'])
def register():
    # Get form data
    first_name = request.form['firstName']
    last_name = request.form['lastName']
    email = request.form['email']
    phone = request.form['phone']
    username = request.form['username']
    password = request.form['password']

    # Get current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Write the registration data to registrations_data.xlsx
    workbook = load_workbook(REGISTRATION_FILE)
    sheet = workbook.active

   # Check if the username already exists
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[5] == username:  # Username is at index 5
            return jsonify({'success': False, 'message': "Username not available."}), 400
        
    

    # Define password requirements
    if not re.match(r'^(?=.*[A-Za-z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$', password):
        return jsonify({'success': False, 'message': "Password must contain at least one letter, one number, and one special character, and be at least 6 characters long."}), 400
    
    sheet.append([timestamp, first_name, last_name, email, phone, username, password])
    workbook.save(REGISTRATION_FILE)

    return jsonify({'success': True, 'message': "Registration successful!"})

@app.route('/contact', methods=['POST'])
def contact():
    # Get form data from the request
    message = request.form['message']

    # Get current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Write contact form data to contact_us.xlsx
    workbook = load_workbook(CONTACT_FILE)
    sheet = workbook.active
    sheet.append([timestamp, message])
    workbook.save(CONTACT_FILE)

    return jsonify({'success': True, 'message': "Your message has been received!"})

@app.route('/report', methods=['POST'])
def report():
    # Handle file upload for pothole reporting
    pincode = request.form['Pin_Code']
    latitude = request.form['latitude']
    longitude = request.form['longitude']

    # Handle file upload
    photo = request.files['photo']
    if photo and photo.filename.endswith('.jpg'):
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], photo.filename)
        photo.save(image_path)
    else:
        return jsonify({'success': False, 'message': "Please upload a JPG file."}), 400

    # Get current timestamp and set status
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    status = "Pending"  # Default status

    # Write the pothole report data to report_potholes.xlsx
    workbook = load_workbook(REPORT_FILE)
    sheet = workbook.active
    sheet.append([timestamp, image_path, pincode, latitude, longitude, status])
    workbook.save(REPORT_FILE)

    user_email = 'tejasviniwagh8@gmail.com'  # Hardcoded user email


    send_report_email( user_email,pincode, latitude, longitude)

    return jsonify({'success': True, 'message': "Pothole report submitted successfully!"})


def send_report_email(user_email, pincode, latitude, longitude):
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    smtp_user = "tanmayzade87@gmail.com"  # Your email address
    smtp_password = "taixfuxzovbxsgoi"  # App-specific password

    msg = MIMEMultipart()
    msg['From'] = smtp_user
    msg['To'] = user_email
    msg['Subject'] = "Pothole Report Submitted"

    body = (f"Dear User,\n\n"
            f"Your pothole report has been submitted successfully.\n"
            f"Location details: Pincode: {pincode}, Latitude: {latitude}, Longitude: {longitude}.\n"
            f"You can view it on Google Maps: https://www.google.com/maps/search/?api=1&query={latitude},{longitude}\n\n"
            "Best regards,\nYour Team")
    
    msg.attach(MIMEText(body, 'plain'))

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
        print("Email sent successfully!")  # Debugging output
        return True
    except Exception as e:
        print(f"Error sending email: {e}")  # Improved error output
        return False



@app.route('/get-coordinates', methods=['GET'])
def get_coordinates():
    if not os.path.exists(REPORT_FILE):
        logging.error("Excel file not found.")
        return jsonify({"error": "Excel file not found"}), 404
    
    # Load the Excel file
    try:
        df = pd.read_excel(REPORT_FILE)
        logging.debug(f"Excel file loaded successfully: {REPORT_FILE}")
    except Exception as e:
        logging.error(f"Error loading Excel file: {str(e)}")
        return jsonify({"error": str(e)}), 500
    
    # Check if required columns exist
    required_columns = ['Timestamp', 'Latitude', 'Longitude']
    if not all(col in df.columns for col in required_columns):
        missing_cols = [col for col in required_columns if col not in df.columns]
        logging.error(f"Missing columns in Excel file: {missing_cols}")
        return jsonify({"error": f"Excel file must contain columns: {required_columns}"}), 400
    
    # Extract coordinates into a list of dictionaries
    coordinates = df[required_columns].dropna()
    
    # Log the raw coordinates before validation
    logging.debug(f"Raw coordinates extracted: {coordinates.to_dict(orient='records')}")

    # Validate the latitude and longitude values
    coordinates = [
        coord for coord in coordinates.to_dict(orient='records') 
        if isinstance(coord['Latitude'], (int, float)) and isinstance(coord['Longitude'], (int, float))
    ]
    
    # Log the validated coordinates
    logging.debug(f"Validated coordinates: {coordinates}")
    
    return jsonify(coordinates)




@app.route('/login', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']

    # Load the Excel file
    workbook = load_workbook(REGISTRATION_FILE)
    sheet = workbook.active

    # Check for matching username and password
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[5] == username and row[6]== password :  # Username at index 5 and Password at index 6
            return jsonify({'success': True, 'message': "Login successful!"})

    return jsonify({'success': False, 'message': "Invalid username or password."})







# Route for retrieving the reports for the admin dashboard
@app.route('/get-reports', methods=['GET'])
def get_reports():
    workbook = load_workbook(REPORT_FILE)
    sheet = workbook.active
    reports = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        report = {
            'Timestamp': row[0],
            'Image Path': row[1],
            'Location': f"Pincode: {row[2]}, Lat: {row[3]}, Long: {row[4]}",

            # 'pincode': row[2],       # Separate field for Pincode
            # 'latitude': row[3],      # Separate field for Latitude
            # 'longitude': row[4],     # Separate field for Longitude
            'Status': row[5]
        }
        reports.append(report)

    return jsonify(reports)

# Route for rendering the admin page
@app.route('/admin')
def admin_page():
    return render_template('admin.html')

@app.route('/update-status', methods=['POST'])
def update_status():
    data = request.get_json()
    index = data['index']
    new_status = data['status']

    # Load the existing reports
    workbook = load_workbook(REPORT_FILE)
    sheet = workbook.active

    # Update the status of the specified report
    row_index = index + 2  # Adjusting for header row
    sheet.cell(row=row_index, column=6).value = new_status
    workbook.save(REPORT_FILE)

    # Hardcoded user email
    user_email = 'tejasviniwagh8@gmail.com'  # Hardcoded user email
    user_name = 'User'  # You may want to retrieve this from the Excel as well

    # Log email information
    print(f"Sending email to: {user_email} with status: {new_status}")

    email_response = send_email(user_email, user_name, new_status)
    print(email_response)  # Debugging output
    return jsonify({
        'success': True,
        'message': "Status updated successfully!",
        'email_status': email_response
    })

# Function to send a confirmation email with the updated status
def send_email(user_email, user_name, new_status):
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    smtp_user = "tanmayzade87@gmail.com"  # Your email address
    smtp_password = "taixfuxzovbxsgoi"  # App-specific password

    msg = MIMEMultipart()
    msg['From'] = smtp_user
    msg['To'] = user_email
    msg['Subject'] = "Status Update Confirmation"

    body = (f"Dear {user_name},\n\n"
            f"Your status has been updated to: {new_status}.\n\n"
            "Best regards,\nYour Healthcare Team")
    
    msg.attach(MIMEText(body, 'plain'))

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
        print("Email sent successfully!")  # Debugging output
        return True
    except Exception as e:
        print(f"Error sending email: {e}")  # Improved error output
        return False
    
@app.route('/get-pothole-data')
def get_pothole_data():
    try:
        # Read the Excel file
        if not os.path.exists(EXCEL_FILE_PATH):
            # Return sample data if the file doesn't exist (for testing)
            return jsonify({
                "error": "Excel file not found",
                "sample_data": True
            }), 404
        
        df = pd.read_excel(EXCEL_FILE_PATH)
        
        # Convert to CSV temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp:
            df.to_csv(tmp.name, index=False)
            tmp_path = tmp.name
        
        # Send the CSV file
        return send_file(
            tmp_path,
            mimetype='text/csv',
            as_attachment=True,
            download_name='pothole_data.csv'
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Alternative route that returns JSON directly
@app.route('/get-pothole-stats')
def get_pothole_stats():
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            # Sample data for testing
            return jsonify({
                "Pending": 15,
                "In Progress": 8,
                "Completed": 12
            })
        
        df = pd.read_excel(EXCEL_FILE_PATH)
        
        # Fill empty Status values with "Pending"
        df['Status'] = df['Status'].fillna('Pending')
        
        # Count occurrences of each status
        status_counts = df['Status'].value_counts().to_dict()
        
        # Ensure all statuses are represented
        for status in ['Pending', 'In Progress', 'Completed']:
            if status not in status_counts:
                status_counts[status] = 0
        
        return jsonify(status_counts)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
